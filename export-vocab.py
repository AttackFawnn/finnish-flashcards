#!/usr/bin/env python3
"""
Finnish Flashcard Vocabulary Updater
Reads Finnish_Vocab_Index.xlsx and updates index.html with new VOCAB_DATA
"""

import openpyxl
import os
import sys

def main():
    # File paths
    excel_file = 'Finnish_Vocab_Index.xlsx'
    index_file = 'index.html'
    
    # Check if Excel file exists
    if not os.path.exists(excel_file):
        print(f"❌ Error: {excel_file} not found in current directory")
        sys.exit(1)
    
    print("📖 Reading Finnish_Vocab_Index.xlsx...")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_file)
        sheet_names = wb.sheetnames
        print(f"   Found sheets: {', '.join(sheet_names)}")
        
        # Process Finnish Vocab sheet
        vocab_sheet = wb['Finnish Vocab']
        vocab_data = []
        
        for row_idx, row in enumerate(vocab_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:  # Skip empty rows
                continue
            
            finnish, english, category, skip = row[0], row[1], row[2], row[3]
            
            # Skip if marked as "Skip?"
            if skip and str(skip).strip() != '':
                continue
            
            vocab_data.append({
                'finnish': str(finnish).strip() if finnish else '',
                'english': str(english).strip() if english else '',
                'category': str(category).strip() if category else ''
            })
        
        print(f"   Finnish Vocab: {len(vocab_data)} words")
        
        # Process Letter Sounds sheet
        letters_sheet = wb['Letter Sounds']
        letter_sounds = []
        
        for row_idx, row in enumerate(letters_sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None or str(row[0]).strip() == 'Letter':  # Skip empty/header
                continue
            
            letter, sound, approx = row[0], row[1], row[2]
            
            if letter and str(letter).strip():
                letter_sounds.append({
                    'letter': str(letter).strip(),
                    'sound': str(sound).strip() if sound else '',
                    'approx': str(approx).strip() if approx else ''
                })
        
        print(f"   Letter Sounds: {len(letter_sounds)} letters")
        
        # Combine all words
        all_words = []
        
        # Add vocabulary
        for word in vocab_data:
            all_words.append(word)
        
        # Add letter sounds
        for letter in letter_sounds:
            all_words.append({
                'finnish': letter['letter'],
                'english': f"{letter['sound']} ({letter['approx']})",
                'category': 'Letter Sounds'
            })
        
        print(f"   Total words: {len(all_words)}")
        
        # Generate JavaScript code
        js_code = 'const VOCAB_DATA = [\n'
        for idx, word in enumerate(all_words):
            finnish = word['finnish'].replace('"', '\\"')
            english = word['english'].replace('"', '\\"')
            category = word['category'].replace('"', '\\"')
            
            js_code += f'            {{"finnish": "{finnish}", "english": "{english}", "category": "{category}"}}'
            js_code += ',' if idx < len(all_words) - 1 else ','
            js_code += '\n'
        js_code += '        ];'
        
        # Read index.html
        print("📝 Updating index.html...")
        
        if not os.path.exists(index_file):
            print(f"❌ Error: {index_file} not found in current directory")
            sys.exit(1)
        
        with open(index_file, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # Find and replace VOCAB_DATA
        import re
        pattern = r'const VOCAB_DATA = \[[\s\S]*?\];(?=\s*function FlashcardApp)'
        
        if not re.search(r'const VOCAB_DATA = \[', html_content):
            print("❌ Error: Could not find VOCAB_DATA in index.html")
            sys.exit(1)
        
        html_content = re.sub(pattern, js_code, html_content)
        
        # Write updated index.html
        with open(index_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print("✅ Success! index.html updated with new vocabulary")
        print(f"   Total: {len(all_words)} words")
        print("")
        print("📤 Next steps:")
        print("   git add .")
        print("   git commit -m \"Update vocabulary\"")
        print("   git push")
        
    except KeyError as e:
        print(f"❌ Error: Could not find sheet {e}. Make sure you have 'Finnish Vocab' and 'Letter Sounds' sheets.")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
